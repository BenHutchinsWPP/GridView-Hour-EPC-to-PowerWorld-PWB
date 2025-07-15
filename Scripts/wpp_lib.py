from pathlib import Path
import os
import multiprocessing as mp
import numpy as np
import pandas as pd
import openpyxl.utils
import win32com.client
import warnings

# Filter warnings on applymap and fillna for now. 
# To Do: Identify a future-proof version of these calls. 
warnings.filterwarnings("ignore", category=FutureWarning, message=".*DataFrame[.]applymap.*")
warnings.filterwarnings("ignore", category=FutureWarning, message=".*fillna.*")

mva_mismatch_threshold = 1.0 

def chk(SimAuto, SimAutoOutput, Message):
    """
    Function used to catch and display errors passed back from SimAuto

    SimAuto return object format:
    [0] = Error message, if any
    [1+] = Return data
    """

    if SimAutoOutput[0] != '':
        print('Error: ' + SimAutoOutput[0])
        return None
    # else:
    #     print(Message)

    if len(SimAutoOutput) == 1:
        return None
    elif len(SimAutoOutput) == 2:
        return SimAutoOutput[1]
    else:
        return SimAutoOutput[1:]

def get_param(SimAuto, table: str, parameters: list[str], filter_group: str = ''):
    msg = 'GetParametersMultipleElementRect(' + table + ': [' + ', '.join(parameters) + '])'
    return_value = chk(SimAuto, SimAuto.GetParametersMultipleElementRect(table, parameters, filter_group), msg)
    return return_value

def get_param_df(SimAuto, table: str, parameter_type: dict[str,type], filter_group: str = '') -> pd.DataFrame:
    # Get data from PowerWorld. 
    parameter_list: list[str] = list(parameter_type.keys())
    rows: list[list[str]] = get_param(SimAuto, table, parameter_list, filter_group)
    # Pack into a dataframe. 
    df = pd.DataFrame(data=rows, columns=parameter_list)
    # Trim all strings. 
    # TO DO: fix future warning. 
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    # Change all data types to the proper types. 
    for parameter in parameter_list:
        if(parameter_type[parameter] in [int, float]):
            # Handle numeric types with coerce for empty strings. 
            df[parameter] = pd.to_numeric(df[parameter], errors='coerce')
        else: 
            df[parameter] = df[parameter].astype(parameter_type[parameter])
    return df

def set_param(SimAuto, table: str, parameters: list[str], rows: list[list[str]]):
    msg = 'ChangeParametersMultipleElementRect(' + table + ': [' + ', '.join(parameters) + '])'
    return_value = chk(SimAuto, SimAuto.ChangeParametersMultipleElementRect(table, parameters, rows), msg)
    return return_value

def set_param_df(SimAuto, table, df: pd.DataFrame):
    df = df.reset_index()
    if(len(df) == 0):
        return ''

    # Get parameters. 
    parameters: list[str] = df.columns.tolist()
    # Convert df into list of lists. All numerical values which are "nan" must be treated as empty strings. 
    # TO DO: fix future warning. 
    rows: list[list[str]] = df.fillna('').astype(str).values.tolist()
    # Set data in PowerWorld. 
    return_value = set_param(SimAuto, table, parameters, rows)
    return return_value

def open_case(SimAuto, fp) -> bool:
    # Attempts to open a case.
    # Error case: message = ('OpenCase: Errors have occurred',)
    # Success case: message = ('',)

    if not Path(fp).exists():
        print(f'Path does not exist: {str(fp)}')
        return False
    
    message = SimAuto.OpenCase(fp)

    if 'OpenCase: Error' in message[0]:
        print(f'Could not open: {str(fp)}')
        return False

    print(f'Opened: {str(fp)}')
    return True

def save_case(SimAuto, fp, case_format = 'PWB22') -> bool:
    # Attempts to save a case.
    # No case in memory to save out?
    #   Error case: message = ('SaveCase: Error trying to save c:\\case 1.pwb - aborted',)
    # Writing to a parent folder which doesn't exist?
    #   Error case: message = ('SaveCase: Windows has prevented us from writing to the file c:\\foobar\\nonexistent directory.pwb. Verify your write privileges and that the specified parent directory exists.',)
    # Success case: message = ('',)

    if not Path(fp).parent.exists():
        print(f'Path does not exist: {str(fp)}')
        return False
    
    message = SimAuto.SaveCase(fp, case_format, True)

    if 'SaveCase: ' in message[0]:
        print(f'Could not save to: {str(fp)}')
        print(message[0])
        return False

    print(f'Saved: {str(fp)}')
    return True

def solve(SimAuto, mva_mismatch_threshold = 1.0) -> bool:
    # Solve.
    SimAuto.RunScriptCommand('EnterMode(RUN);')
    result = SimAuto.RunScriptCommand('SolvePowerFlow(RECTNEWT);')

    # Error string. Return early with False if it didn't solve. 
    if result[0] != '': 
        print(result[0])
        return False
    SimAuto.RunScriptCommand('EnterMode(EDIT);')

    # Get mismatch. 
    df = get_param_df(SimAuto, 'Bus', {'Busnum':int, 'MismatchP':float, 'MismatchQ':float})
    df['MismatchS'] = (df['MismatchP']**2.0 + df['MismatchQ']**2.0)**0.5
    max_mismatch = df['MismatchS'].abs().max()

    # print(f'Max Mismatch (S) = {max_mismatch}')

    return max_mismatch < mva_mismatch_threshold

# Given:
#   SimAuto: PowerWorld SimulatorAuto object
#   df: A dataframe of object parameters to set. 
# Attempt to apply all the changes in bulk edits using set_param_df().
# If it fails, revert to the previous system state, and bifurcate the df into two halves to test recursively until reaching 1 item. 
# If individual items fail, keep track of those failures on a row-by-row basis.
def set_param_df_recursive(SimAuto, table: str, df: pd.DataFrame):
    SimAuto.SaveState()

    # Attempt all the edits in a single bulk step. 
    message = set_param_df(SimAuto, table, df)

    if not solve(SimAuto, mva_mismatch_threshold):
        SimAuto.LoadState()
        # Failed to do all changes at once! Revert, and try individual branch changes. 
        print(f'Failed to set parameters on all elements at the same time. Bifurcating into half the list size.')

        # If dataframe is already only 1 item, record the failure on that row of the slice of the dataframe. 
        # 'ExclusionReason' = 'Diverged'. 
        # 'Include' = False.
        if len(df) == 1:
            df.at[df.index[0], 'Include'] = False
            df.at[df.index[0], 'ExclusionReason'] = 'Diverged'
            return
        
        # Split the list into two halves to process. 
        set_param_df_recursive(SimAuto, table, df.iloc[:len(df)//2])
        set_param_df_recursive(SimAuto, table, df.iloc[len(df)//2:])

    return 

def get_case_data(SimAuto) -> dict[str,dict[str,object]]:
    print('get_case_data()')
    
    # Define all the required network-parameters to fully define each object. 
    bus_params: dict[str,type] = {
        'ObjectID': str
        ,'Number': int
        ,'Name': str
        ,'NomkV': float
        ,'Slack': str
        ,'NomB': float
        ,'NomG': float
        ,'Vpu': float
        ,'Vangle': float
        ,'DCLossMultiplier': float
        ,'AreaNumber': int
        ,'ZoneNumber': int
        ,'BANumber': int
        ,'OwnerNumber': int
        ,'SubNumber': int
        ,'Monitor': str
        ,'LimitSet': str
        ,'UseSpecificLimits': str
        ,'LimitLowA': float
        ,'LimitLowB': float
        ,'LimitLowC': float
        ,'LimitLowD': float
        ,'LimitHighA': float
        ,'LimitHighB': float
        ,'LimitHighC': float
        ,'LimitHighD': float
        ,'Latitude': float
        ,'Longitude': float
        ,'TopologyBusType': str
        ,'Priority': float
        ,'EMSType': str
        ,'EMSID': str
        ,'DataMaintainerAssign': str
        ,'DataMaintainerInherit': str
        ,'DataMaintainerInheritBlock': str
        ,'AllLabels': str
    }

    load_params: dict[str,type] = {
        'ObjectID': str
        ,'BusNum': int
        ,'ID': str
        ,'Status': str
        ,'AGC': str
        ,'SMW': float
        ,'SMvar': float
        ,'IMW': float
        ,'IMvar': float
        ,'ZMW': float
        ,'ZMvar': float
        ,'DistStatus': str
        ,'DistMWInput': float
        ,'DistMvarInput': float
        ,'Interruptible': str
        ,'MWMax': float
        ,'MWMin': float
        ,'DistMWMax': float
        ,'DistMWMin': float
        ,'DistUnitTypeCode': str
        ,'LoadModelGroup': str
        ,'AreaNumber': int
        ,'ZoneNumber': int
        ,'BANumber': int
        ,'OwnerNumber': int
        ,'EMSType': str
        ,'EMSID': str
        ,'DataMaintainerAssign': str
        ,'DataMaintainerInherit': str
        ,'AllLabels': str
    }

    gen_params: dict[str,type] = {
        # Note: ReactiveCapability curve data is not necessary in the comparisons.
        # The Anchor Data Set may have additional generators, but none of them use ReactiveCapability curves.
        'ObjectID': str
        ,'BusNum': int
        ,'ID': str
        ,'Status': str
        ,'VoltSet': float
        ,'VoltSetTol': float
        ,'RegBusNum': int
        ,'RegFactor': float
        ,'AGC': str
        ,'PartFact': float
        ,'MWSetPoint': float
        ,'MWMax': float
        ,'MWMin': float
        ,'EnforceMWLimit': str
        ,'AVR': str
        ,'MvarSetPoint': float
        ,'MvarMax': float
        ,'MvarMin': float
        ,'UseCapCurve': str
        ,'WindContMode': str
        ,'WindContModePF': float
        ,'UseLineDrop': str
        ,'Rcomp': float
        ,'Xcomp': float
        ,'VoltageDroopControl': str
        ,'MVABase': float
        ,'GenR': float
        ,'GenX': float
        ,'StepR': float
        ,'StepX': float
        ,'StepTap': float
        ,'GovRespLimit': str
        ,'UnitTypeCode': str
        ,'FuelTypeCode': str
        ,'AreaNumber': int
        ,'ZoneNumber': int
        ,'BANumber': int
        ,'OwnerNum1': int
        ,'OwnerPerc1': float
        ,'OwnerNum2': int
        ,'OwnerPerc2': float
        ,'OwnerNum3': int
        ,'OwnerPerc3': float
        ,'OwnerNum4': int
        ,'OwnerPerc4': float
        ,'OwnerNum5': int
        ,'OwnerPerc5': float
        ,'OwnerNum6': int
        ,'OwnerPerc6': float
        ,'OwnerNum7': int
        ,'OwnerPerc7': float
        ,'OwnerNum8': int
        ,'OwnerPerc8': float
        ,'EMSType': str
        ,'EMSID': str
        ,'DataMaintainerAssign': str
        ,'DataMaintainerInherit': str
        ,'AllLabels': str
    }

    branch_params: dict[str,type] = {
        # Non-transformer branches.
        'ObjectID': str
        ,'BusNumFrom': int
        ,'BusNumTo': int
        ,'Circuit': str
        ,'BranchDeviceType': str
        ,'ConsolidateAllow': str
        ,'OpenOrCloseBreakersAllow': str
        ,'Status': str
        ,'StatusNormal': str
        ,'ByPass': str
        ,'MeteredBus': str
        ,'R': float
        ,'X': float
        ,'B': float
        ,'G': float
        ,'LineLength': float
        ,'Monitor': str
        ,'LimitSe': str
        ,'LimitMVAA': float
        ,'LimitMVAB': float
        ,'LimitMVAC': float
        ,'LimitMVAD': float
        ,'LimitMVAE': float
        ,'LimitMVAF': float
        ,'LimitMVAG': float
        ,'LimitMVAH': float
        ,'LimitMVAI': float
        ,'LimitMVAJ': float
        ,'LimitMVAK': float
        ,'LimitMVAL': float
        ,'LimitMVAM': float
        ,'LimitMVAN': float
        ,'LimitMVAO': float
        ,'OwnerNum1': int
        ,'OwnerPerc1': float
        ,'OwnerNum2': int
        ,'OwnerPerc2': float
        ,'OwnerNum3': int
        ,'OwnerPerc3': float
        ,'OwnerNum4': int
        ,'OwnerPerc4': float
        ,'OwnerNum5': int
        ,'OwnerPerc5': float
        ,'OwnerNum6': int
        ,'OwnerPerc6': float
        ,'OwnerNum7': int
        ,'OwnerPerc7': float
        ,'OwnerNum8': int
        ,'OwnerPerc8': float
        ,'EMSType': str
        ,'EMSID': str
        ,'EMSLineID': str
        ,'EMSCBTyp': str
        ,'EMSID2From': str
        ,'EMSID2To': str
        ,'DataMaintainerAssign': str
        ,'DataMaintainerInherit': str
        ,'AllLabels': str
    }

    transformer_params: dict[str,type] = {
        # Transformers in the branch table.
        'ObjectID': str
        ,'BusNumFrom': int
        ,'BusNumTo': int
        ,'Circuit': str
        ,'BranchDeviceType': str
        ,'Status': str
        ,'StatusNormal': str
        ,'ByPass': str
        ,'MeteredBus': str
        ,'ControlType': str
        ,'AutoControl': str
        ,'RegBusNum': int
        ,'UseLineDrop': str
        ,'Rcomp': float
        ,'Xcomp': float
        ,'RegMax': float
        ,'RegMin': float
        ,'RegTargetType': str
        ,'XFMVABase': float
        ,'XFNomkVbaseFrom': float
        ,'XFNomkVbaseTo': float
        ,'Rxfbase': float
        ,'Xxfbase': float
        ,'Gxfbase': float
        ,'Bxfbase': float
        ,'Gmagxfbase': float
        ,'Bmagxfbase': float
        ,'TapFixedFrom': float
        ,'TapFixedTo': float
        ,'TapMaxxfbase': float
        ,'TapMinxfbase': float
        ,'TapStepSizexfbase': float
        ,'Tapxfbase': float
        ,'Phase': float
        ,'ImpCorrTable': float
        ,'LineLength': float
        ,'Monitor': str
        ,'LimitSet': str
        ,'LimitMVAA': float
        ,'LimitMVAB': float
        ,'LimitMVAC': float
        ,'LimitMVAD': float
        ,'LimitMVAE': float
        ,'LimitMVAF': float
        ,'LimitMVAG': float
        ,'LimitMVAH': float
        ,'LimitMVAI': float
        ,'LimitMVAJ': float
        ,'LimitMVAK': float
        ,'LimitMVAL': float
        ,'LimitMVAM': float
        ,'LimitMVAN': float
        ,'LimitMVAO': float
        ,'OwnerNum1': int
        ,'OwnerPerc1': float
        ,'OwnerNum2': int
        ,'OwnerPerc2': float
        ,'OwnerNum3': int
        ,'OwnerPerc3': float
        ,'OwnerNum4': int
        ,'OwnerPerc4': float
        ,'OwnerNum5': int
        ,'OwnerPerc5': float
        ,'OwnerNum6': int
        ,'OwnerPerc6': float
        ,'OwnerNum7': int
        ,'OwnerPerc7': float
        ,'OwnerNum8': int
        ,'OwnerPerc8': float
        ,'EMSType': str
        ,'EMSID': str
        ,'EMSLineID': str
        ,'EMSCBTyp': str
        ,'EMSID2From': str
        ,'EMSID2To': str
        ,'DataMaintainerAssign': str
        ,'DataMaintainerInherit': str
        ,'AllLabels': str
    }

    lineshunt_params: dict[str,type] = {
        'ObjectID': str
        ,'BusNumFrom': int
        ,'BusNumTo': int
        ,'Circuit': str
        ,'ID': str
        ,'BusNumLoc': int
        ,'Status': str
        ,'MWNom': float
        ,'MvarNom': float
        ,'OwnerNum1': int
        ,'OwnerPerc1': float
        ,'OwnerNum2': int
        ,'OwnerPerc2': float
        ,'OwnerNum3': int
        ,'OwnerPerc3': float
        ,'OwnerNum4': int
        ,'OwnerPerc4': float
        ,'DataMaintainerAssign': str
        ,'DataMaintainerInherit': str
    }

    multisectionline_params: dict[str,type] = {
        'ObjectID': str
        ,'BusNumFrom': int
        ,'BusNumTo': int
        ,'Circuit': str
        ,'AllowMixedStatus': str
        ,'BusInt:0': int
        ,'BusInt:1': int
        ,'BusInt:2': int
        ,'BusInt:3': int
        ,'BusInt:4': int
        ,'BusInt:5': int
        ,'BusInt:6': int
        ,'BusInt:7': int
        ,'BusInt:8': int
        ,'BusInt:9': int
        ,'BusInt:10': int
        ,'BusInt:11': int
        ,'BusInt:12': int
        ,'BusInt:13': int
        ,'BusInt:14': int
        ,'BusInt:15': int
        ,'BusInt:16': int
        ,'BusInt:17': int
        ,'BusInt:18': int
        ,'BusInt:19': int
        ,'BusInt:20': int
        ,'DataMaintainerAssign': str
    }

    case_dict = {
        'Bus':{
            'table_name': 'Bus'
            ,'df': get_param_df(SimAuto, 'Bus', bus_params)
        }
        ,'Load':{
            'table_name': 'Load'
            ,'df': get_param_df(SimAuto, 'Load', load_params)
        }
        ,'Gen':{
            'table_name': 'Gen'
            ,'df': get_param_df(SimAuto, 'Gen', gen_params)
        }
        ,'Branch':{
            'table_name': 'Branch'
            ,'df': get_param_df(SimAuto, 'Branch', branch_params, "BranchDeviceType notcontains 'Transformer'")
        }
        ,'Transformer':{
            'table_name': 'Branch'
            ,'df': get_param_df(SimAuto, 'Branch', transformer_params, "BranchDeviceType = 'Transformer'")
        }
        ,'LineShunt':{
            'table_name': 'LineShunt'
            ,'df': get_param_df(SimAuto, 'LineShunt', lineshunt_params)
        }
        # ,'MultiSectionLine':{
        #     'table_name': 'MultiSectionLine'
        #     ,'df': get_param_df(SimAuto, 'MultiSectionLine', multisectionline_params)
        # }
    }

    return case_dict

def create_dummy_bus_aux(SimAuto, dummy_bus_fp: Path):
    command_str = 'SaveData("'+str(dummy_bus_fp)+'", AUX, MultiSectionLine, [BusNumFrom,BusNameFrom,BusNumTo,BusNameTo,Circuit], [Bus], , [], NO, NO);'
    retVal = SimAuto.RunScriptCommand(command_str)
    print(retVal)

    content = dummy_bus_fp.read_text(encoding="utf-8")
    content = content.replace("&", " ")
    content = content.replace("<SUBDATA Bus>", "<SUBDATA BusRenumber>")
    dummy_bus_fp.write_text(content, encoding="utf-8")

    return

def create_missing_elements(SimAuto, left_case_dict, right_case_dict) -> dict[str,pd.DataFrame]:
    # Any elements on the left, which aren't on the right, will get created. 
    print('create_missing_elements()')
    
    SimAuto.RunScriptCommand('EnterMode(EDIT);')
    SimAuto.CreateIfNotFound = True

    missing_dict: dict[str,pd.DataFrame] = {}

    for element_type in left_case_dict:
        table_name = left_case_dict[element_type]['table_name']
        print(f'Creating missing {element_type} elements')

        left_df = left_case_dict[element_type]['df']
        right_df = right_case_dict[element_type]['df']

        missing_df = left_df[~left_df['ObjectID'].isin(right_df['ObjectID'])].copy(deep=True)

        missing_dict[element_type] = missing_df

        # To create objects, you cannot have 'ObjectID' in the list of fields. 
        missing_df = missing_df.drop(columns=['ObjectID'])

        # Insert the elements into the case with an initially open status. 
        if 'Status' in left_df.columns:
            missing_df['Status'] = 'Open'

        # Several GridView transformers are in automatic control with crazy tap settings. Disable automatic controls.
        if element_type == 'Transformer':
            missing_df['AutoControl'] = 'NO'

        message = set_param_df(SimAuto, table_name, missing_df)

    return missing_dict

def create_giant_swing(SimAuto, fault_df) -> pd.DataFrame:
    """
    Creates a giant swing/slack unit on the bus with the max MVA fault duty. 
    """
    # Get the row with the maximum 'MVA' value
    max_mva_busnum = fault_df.loc[fault_df['MVA'].idxmax(), 'BusNumber']

    # Get current bus information at the max MVA bus. 
    bus_params: dict[str,type] = {
        'Number': int
        ,'Name': str
        ,'NomkV': float
        ,'Vpu': float
    }
    bus_df = get_param_df(SimAuto, 'Bus', bus_params)

    bus_row = bus_df.loc[bus_df['Number'] == max_mva_busnum]
    name = bus_row['Name'].iloc[0]
    nomkV = bus_row['NomkV'].iloc[0]
    vpu = bus_row['Vpu'].iloc[0]

    # Create the giant swing. 
    swing_dict = {
        'BusNum': max_mva_busnum
        ,'BusName': name
        ,'NomkV': nomkV
        ,'ID': "xx"
        ,'VoltSet': vpu
        ,'MW': 0.00
        ,'Status': "Closed"
        ,'AVR': "YES"
        ,'AGC': "YES"
        ,'MvarMax': 6573.70
        ,'MvarMin': -6573.70
        ,'Mvar': 0.00
        ,'MWMax': 20000.00
        ,'MWMin': -20000.00
        ,'FuelTypeCode': "MWH"
        ,'UnitTypeEPC': "42"
        ,'Memo': "Giant Swing for Initial Balance"
    }
    swing_df = pd.DataFrame.from_dict([swing_dict])

    SimAuto.RunScriptCommand('EnterMode(EDIT);')
    SimAuto.CreateIfNotFound = True
    set_param_df(SimAuto, 'Gen',swing_df)

    # Disable any other swings. 
    SimAuto.RunScriptCommand('SetData(Bus,[Slack],[NO],ALL);')

    # Set current slack as system swing. 
    SimAuto.RunScriptCommand(f'SetData(Bus,[Number,Slack],[{int(max_mva_busnum)}, YES]);')

    # Turn off MW AGC
    settings_df = pd.DataFrame({
       'Option':['ChkMWAGC']
       ,'Value':['NO']
    })
    set_param_df(SimAuto, 'Sim_Solution_Options_Value', settings_df)

    return swing_df

def create_distgen_XN_loads(SimAuto, gv_fps: Path, toposeed_fp: Path) -> pd.DataFrame:
    """
    Gathers load data from all GridView EPCs
        (since they are dynamically generated by GridView for each hour)
    Opens the TopoSeed case. Creates all X1/X2/X3 etc distributed generation loads which don't exist already.
    New loads will be in a normal-open status, with MW=0 MVAR=0 for all related values. 
    Returns a dataframe of all distributed generation loads which were created. 
    """
    # Open each case in gv_fps, and get the load data from get_case_data()
    gv_load_df_list = []
    for gv_fp in gv_fps:
        open_case(SimAuto, gv_fp)
        case_dict = get_case_data(SimAuto)
        gv_load_df = case_dict['Load']['df']
        SimAuto.CloseCase()
        gv_load_df_list.append(gv_load_df)
    
    # Merge all load dataframes into one. 
    gv_load_df = pd.concat(gv_load_df_list, ignore_index=True)
    # Remove duplicates on ObjectID.
    gv_load_df = gv_load_df.drop_duplicates(subset='ObjectID', keep='first')
    
    # Get the topology seed data. 
    open_case(SimAuto, toposeed_fp)
    case_dict = get_case_data(SimAuto)
    pw_load_df = case_dict['Load']['df']

    # Get all gv_load_df rows which do not yet exist in pw_load_df.
    missing_df = gv_load_df[~gv_load_df['ObjectID'].isin(pw_load_df['ObjectID'])].copy(deep=True)
    missing_df['Status'] = 'Open'
    missing_df['SMW'] = 0
    missing_df['SMvar'] = 0
    missing_df['DistStatus'] = 'Open'
    missing_df['DistMWInput'] = 0
    missing_df['DistMvarInput'] = 0
    missing_df['IMW'] = 0
    missing_df['IMvar'] = 0
    missing_df['ZMW'] = 0
    missing_df['ZMvar'] = 0

    # Create the missing loads.
    # To create objects, you cannot have 'ObjectID' in the list of fields. 
    missing_df = missing_df.drop(columns=['ObjectID'])
    SimAuto.RunScriptCommand('EnterMode(EDIT);')
    SimAuto.CreateIfNotFound = True
    message = set_param_df(SimAuto, 'Load', missing_df)
    SimAuto.RunScriptCommand('EnterMode(RUN);')

    return missing_df

def compute_pw_targets(SimAuto, left_fp: Path, right_fp: Path) -> list[pd.DataFrame]:
    """
    Returns gen & load dataframes. 
    Left case: The load & generation you wish to have (Target). 
    Right case: The case you are using, which should take in the gen & load from the Left case. 
    "MWSetPoint": Right Case Value. 
    "MWSetPoint_Target": Left Case Value. 
    """

    gen_params: dict[str,type] = {
        'ObjectID': str
        ,'BusNum': int
        ,'BusName': str
        ,'NomkV': float
        ,'ID': str
        ,'Status': str
        ,'MWSetPoint': float
    }

    load_params: dict[str,type] = {
        'ObjectID': str
        ,'BusNum': int
        ,'BusName': str
        ,'NomkV': float
        ,'ID': str
        ,'Status': str
        ,'SMW': float
        ,'SMvar': float
        ,'DistStatus': str
        ,'DistMWInput': float
        ,'DistMvarInput': float
        # Not frequently used: 
        # ,'IMW': float
        # ,'IMvar': float
        # ,'ZMW': float
        # ,'ZMvar': float
    }

    # Get data from left case.
    if not open_case(SimAuto, left_fp):
        raise
    left_gen_df = get_param_df(SimAuto, 'Gen', gen_params)
    left_load_df = get_param_df(SimAuto, 'Load', load_params)

    # Get data from right case.
    if not open_case(SimAuto, right_fp):
        raise
    right_gen_df = get_param_df(SimAuto, 'Gen', gen_params)
    right_load_df = get_param_df(SimAuto, 'Load', load_params)

    # Put current case values, and the "Target", side by side. 
    gen_target_df = right_gen_df.merge(
        left_gen_df
        ,on='ObjectID'
        ,how='left'
        ,suffixes=('', '_Target')
    )
    gen_target_df.drop(columns=['BusNum_Target', 'BusName_Target', 'NomkV_Target', 'ID_Target'], inplace=True)

    # If the generator exists on the right, but not on the left (target), set to 0 MW and Open it.
    gen_target_df.fillna({
        'Status_Target': "Open"
        ,'MWSetPoint_Target': 0
    }, inplace=True)

    # If the generator is open, set the MW to 0 MW.
    gen_target_df.loc[gen_target_df['Status'] == "Open", 'MWSetPoint'] = 0

    # If the generator target is open, set the MW target to 0 MW.
    gen_target_df.loc[gen_target_df['Status_Target'] == "Open", 'MWSetPoint_Target'] = 0

    load_target_df = right_load_df.merge(
        left_load_df
        ,on='ObjectID'
        ,how='left'
        ,suffixes=('', '_Target')
    )
    load_target_df.drop(columns=['BusNum_Target', 'BusName_Target', 'NomkV_Target', 'ID_Target'], inplace=True)

    load_target_df['Status_Target'] = load_target_df['Status_Target'].fillna("Open")
    load_target_df['DistStatus_Target'] = load_target_df['DistStatus_Target'].fillna("Open")
    load_target_df = load_target_df.fillna(0)

    # If the load target is open, set all targets to 0 MW and 0 MVAR.
    cols_to_zero = ['SMW_Target', 'SMvar_Target', 'DistMWInput_Target', 'DistMvarInput_Target']
    load_target_df.loc[load_target_df['Status_Target'] == "Open", cols_to_zero] = 0

    # If the distributed generation load target is open, set dist targets to 0 MW and 0 MVAR.
    cols_to_zero = ['DistMWInput_Target', 'DistMvarInput_Target']
    load_target_df.loc[load_target_df['DistStatus_Target'] == "Open", cols_to_zero] = 0

    # If the load is open in the base case, set all values to 0 MW and 0 MVAR.
    cols_to_zero = ['SMW', 'SMvar', 'DistMWInput', 'DistMvarInput']
    load_target_df.loc[load_target_df['Status'] == "Open", cols_to_zero] = 0

    # If the distributed generation load is open in the base case, set the starting MW and MVAR to 0. 
    cols_to_zero = ['DistMWInput', 'DistMvarInput']
    load_target_df.loc[load_target_df['DistStatus'] == "Open", cols_to_zero] = 0

    gen_target_df = gen_target_df.set_index('ObjectID')
    load_target_df = load_target_df.set_index('ObjectID')

    # Include all gens and loads in scaling, unless there's a reason to exclude them. 
    gen_target_df['Include'] = True
    load_target_df['Include'] = True
    gen_target_df['ExclusionReason'] = ''
    load_target_df['ExclusionReason'] = ''

    SimAuto.CloseCase()

    return [gen_target_df, load_target_df]

def test_gen_targets(pw_fp: Path, gen_target_df):
    """
    Tests each change individually, and reports which individual changes are not possible. 
    """

    SimAuto = win32com.client.Dispatch("pwrworld.SimulatorAuto")
    open_case(SimAuto, pw_fp)
    solve(SimAuto)
    SimAuto.SaveState()

    # Save previous status and setpoint. 
    gen_target_df['Status_Old'] = gen_target_df['Status']
    gen_target_df['MWSetPoint_Old'] = gen_target_df['MWSetPoint']

    # Test each generation change.
    gen_target_df['Status'] = gen_target_df['Status_Target']
    gen_target_df['MWSetPoint'] = gen_target_df['MWSetPoint_Target']
    gen_target_df['MWSetPoint'].fillna(0, inplace=True)
    gen_target_df['Status'].fillna('Open', inplace=True)

    for i in range(len(gen_target_df)):
        row_df = gen_target_df.iloc[[i]]
        # Set case to target value for this specific element. 
        set_param_df(SimAuto, 'Gen', row_df)
        success = solve(SimAuto)
        gen_target_df.loc[gen_target_df.index[i], 'Success'] = success
        SimAuto.LoadState()

    # Restore previous status and setpoint. 
    gen_target_df['Status'] = gen_target_df['Status_Old']
    gen_target_df['MWSetPoint'] = gen_target_df['MWSetPoint_Old']
    gen_target_df.drop(columns=['Status_Old','MWSetPoint_Old'], inplace=True)

    SimAuto.CloseCase()
    SimAuto = None

    return gen_target_df

def test_gen_targets_parallel(pw_fp: Path, gen_target_df):
    """
    Taking a set of target MW & Status values for generators, tests to see if each one will solve individually.
    """
    num_cores = mp.cpu_count()
    df_splits = np.array_split(gen_target_df, num_cores)

    # Run in parallel:
    with mp.Pool(processes=num_cores) as pool:
        results = pool.starmap(test_gen_targets, [(pw_fp, part) for part in df_splits])

    # Run in series (for debugging):
    # results = [test_gen_targets(pw_fp, part) for part in df_splits]

    gen_target_df = pd.concat(results, ignore_index=True)
    gen_target_df.sort_values(by='Success', ascending=True, inplace=True)
    return gen_target_df

def report_gen_load_balance(gen_target_df, load_target_df):
    # This function will write out a tabular report of the total gen/load MW and differences. 
    # Rows: 
    #   - Gen MW
    #   - Dist Gen MW (In Load Table)
    #   - Load MW
    #   - Total Delta (Gen + Dist Gen - Load)
    # Columns: 
    #   - Base Case
    #   - Target Case
    #   - Delta
    
    base_gen_mw = gen_target_df['MWSetPoint'].sum()
    base_dist_mw = load_target_df['DistMWInput'].sum()
    base_load_mw = load_target_df['SMW'].sum()
    base_total_delta = base_gen_mw + base_dist_mw - base_load_mw

    target_gen_mw = gen_target_df['MWSetPoint_Target'].sum()
    target_dist_mw = load_target_df['DistMWInput_Target'].sum()
    target_load_mw = load_target_df['SMW_Target'].sum()
    target_total_delta = target_gen_mw + target_dist_mw - target_load_mw

    total_delta_df = pd.DataFrame({
        'Base Case': [base_gen_mw, base_dist_mw, base_load_mw, base_total_delta]
        ,'Target Case': [target_gen_mw, target_dist_mw, target_load_mw, target_total_delta]
    })

    total_delta_df.index = ['Gen MW', 'Dist Gen MW', 'Load MW', 'Gen + Dist Gen - Load']

    # Display the report.
    print(total_delta_df.round(0))

    if((target_total_delta / base_total_delta) > 1.5):
        print('-----------------------------------------------------------------------------------')
        print('WARNING: Gen/Load imbalance in the target case is high. This may cause instability.')
        print('-----------------------------------------------------------------------------------')

    return

def iterate_to_gen_load_targets(SimAuto, gen_target_df, load_target_df, pvqv_df, iterations=100):

    def compute_pvqv_exclusions(delta_v_limit = 0.1):
        """
        Checks if the proposed delta-P / delta-Q would cause a linear V change more than the limit.
        Excludes those buses from scaling of gen/load.
        """
        excluded_buses: list = []

        # Check Gen delta-voltage based on delta-MW
        df = gen_target_df.copy()
        df['dp'] = df['MWSetPoint_Target'] - df['MWSetPoint']
        sum_df = df.groupby('BusNum', as_index=False)['dp'].sum()
        merged_df = pvqv_df[['Number', 'Name', 'NomkV', 'SensdVdPself', 'SensdVdQself']].merge(
            sum_df[['BusNum', 'dp']], 
            left_on='Number', 
            right_on='BusNum', 
            how='inner'  # Use 'left', 'right', or 'outer' if needed
            )
        merged_df.drop(columns=['BusNum'], inplace=True) # Remove repeated bus number column. 
        merged_df['dv'] = merged_df['SensdVdPself'] * merged_df['dp']
        merged_df.sort_values(by='dv', key=abs, ascending=False, inplace=True)
        gen_pvqv_df = merged_df[merged_df['dv'].abs() > delta_v_limit]
        excluded_buses.extend(gen_pvqv_df['Number'].tolist())

        # Check Load delta-voltage based on delta-MW & delta-MVAR
        df = load_target_df.copy()
        df['dp'] = df['SMW_Target'] - df['SMW']
        df['dq'] = df['SMvar_Target'] - df['SMvar']
        sum_df = df.groupby('BusNum', as_index=False)[['dp', 'dq']].sum()
        merged_df = pvqv_df[['Number', 'Name', 'NomkV', 'SensdVdPself', 'SensdVdQself']].merge(
            sum_df[['BusNum', 'dp', 'dq']], 
            left_on='Number', 
            right_on='BusNum', 
            how='inner'  # Use 'left', 'right', or 'outer' if needed
            )
        merged_df.drop(columns=['BusNum'], inplace=True) # Remove repeated bus number column. 
        merged_df['dv'] = merged_df['SensdVdPself'] * merged_df['dp'] + merged_df['SensdVdQself'] * merged_df['dq']
        merged_df.sort_values(by='dv', key=abs, ascending=False, inplace=True)
        load_pvqv_df = merged_df[merged_df['dv'].abs() > delta_v_limit]
        excluded_buses.extend(load_pvqv_df['Number'].tolist())

        # Apply exclusions. 
        gen_target_df.loc[gen_target_df['BusNum'].isin(excluded_buses), 'Include'] = False
        gen_target_df.loc[gen_target_df['BusNum'].isin(excluded_buses), 'ExclusionReason'] = 'PVQV'
        load_target_df.loc[load_target_df['BusNum'].isin(excluded_buses), 'Include'] = False
        load_target_df.loc[load_target_df['BusNum'].isin(excluded_buses), 'ExclusionReason'] = 'PVQV'

        # Return PQVQ exclusion calculations. 
        return [gen_pvqv_df, load_pvqv_df]

    def compute_voltage_exclusions(v_min=0.88, v_max=1.12):
        # Once a bus exceeds the vmin/vmax limits, take it out of the scaling equation. 
        # This ensures no voltage collapse related issues occur during scaling. 
        bus_params: dict[str,type] = {
            'Number': int
            ,'Vpu': float
        }
        bus_df = get_param_df(SimAuto, 'Bus', bus_params)
        vpu_dict = bus_df.set_index('Number')['Vpu'].to_dict()
        def item_inclusion(row):
            return (row['Include'] and 
                    vpu_dict[row['BusNum']] > v_min and
                    vpu_dict[row['BusNum']] < v_max
                    )
        gen_target_df['Include'] = gen_target_df.apply(item_inclusion, axis=1)
        load_target_df['Include'] = load_target_df.apply(item_inclusion, axis=1)

        gen_target_df.loc[(gen_target_df['ExclusionReason'] == '') & (gen_target_df['Include'] == False), 'ExclusionReason'] = 'Voltage'
        load_target_df.loc[(load_target_df['ExclusionReason'] == '') & (load_target_df['Include'] == False), 'ExclusionReason'] = 'Voltage'

        return
    
    def close_all_related_gen_load():
        print('close_all_related_gen_load()')
        SimAuto.SaveState()

        print('Closing all related generation at 0 MW output.')
        gens_to_close = (gen_target_df['Status']=='Open') & (gen_target_df['Status_Target']=='Closed')
        gen_target_df.loc[gens_to_close, 'MWSetPoint'] = 0
        gen_target_df.loc[gens_to_close, 'Status'] = 'Closed'
        set_param_df_recursive(SimAuto, 'Gen', gen_target_df)
        if solve(SimAuto):
            SimAuto.SaveState()
        else:
            SimAuto.LoadState()
            print('WARNING: Failed to close all related generators at 0 MW output levels. Check target Excel sheet, and manually check if you can close those generators at 0 MW without divergence. Rolling back change.')
            return
        
        print('Closing all related load at 0 MW // 0 MVAR.')
        loads_to_close = (load_target_df['Status']=='Open') & (load_target_df['Status_Target']=='Closed')
        columns_to_zero = ['SMW', 'SMvar']
        load_target_df.loc[loads_to_close, columns_to_zero] = 0
        load_target_df.loc[loads_to_close, 'Status'] = 'Closed'

        dist_loads_to_close = (load_target_df['DistStatus']=='Open') & (load_target_df['DistStatus_Target']=='Closed')
        columns_to_zero = ['DistMWInput', 'DistMvarInput']
        load_target_df.loc[loads_to_close, columns_to_zero] = 0
        load_target_df.loc[dist_loads_to_close, 'DistStatus'] = 'Closed'

        set_param_df(SimAuto, 'Load', load_target_df)
        if solve(SimAuto):
            SimAuto.SaveState()
        else:
            print('WARNING: Failed to close all related loads at 0 MW output levels. Check target Excel sheet, and manually check if you can close those loads at 0 MW without divergence. Rolling back change.')
            SimAuto.LoadState()

        return

    def set_gen_load_status():
        SimAuto.SaveState()
        print('set_gen_load_status()')
        print('Setting all load statuses...')

        # At this stage, all loads which are intended to be opened should have reached 0 MW and 0 MVAR. 
        # Opening these should have no impact on the model solution.
        load_target_df['Status'] = load_target_df['Status_Target']
        load_target_df['DistStatus'] = load_target_df['DistStatus_Target']
        set_param_df_recursive(SimAuto, 'Load', load_target_df)
        
        # Iterate through each generator, and attempt to change the status to the final target status.
        # This means opening several generators at 0 MW, but they may be providing VAR support. 
        # If it cannot be opened, make note of it and proceed. 
        print('Setting all gen statuses...')
        df = gen_target_df[gen_target_df['Status'] != gen_target_df['Status_Target']].copy()
        df['Status_Old'] = df['Status']
        df['Status'] = df['Status_Target']
        df['FinalStatusChange'] = 'Okay'
        set_param_df_recursive(SimAuto, 'Gen', df)
        df.drop(columns=['Status'], inplace=True)
        return df

    def compute_deltas():
        gen_target_df['MWSetPoint_Delta'] = (gen_target_df['MWSetPoint_Target'] - gen_target_df['MWSetPoint']) / iterations
        load_target_df['SMW_Delta'] = (load_target_df['SMW_Target'] - load_target_df['SMW']) / iterations
        load_target_df['SMvar_Delta'] = (load_target_df['SMvar_Target'] - load_target_df['SMvar']) / iterations
        load_target_df['DistMWInput_Delta'] = (load_target_df['DistMWInput_Target'] - load_target_df['DistMWInput']) / iterations
        load_target_df['DistMvarInput_Delta'] = (load_target_df['DistMvarInput_Target'] - load_target_df['DistMvarInput']) / iterations
        return

    def increment(delta_multiplier = 1.0):
        gen_target_df.loc[gen_target_df['Include']==True, 'MWSetPoint'] += gen_target_df['MWSetPoint_Delta'] * delta_multiplier
        load_target_df.loc[load_target_df['Include']==True, 'SMW'] += load_target_df['SMW_Delta'] * delta_multiplier
        load_target_df.loc[load_target_df['Include']==True, 'SMvar'] += load_target_df['SMvar_Delta'] * delta_multiplier
        load_target_df.loc[load_target_df['Include']==True, 'DistMWInput'] += load_target_df['DistMWInput_Delta'] * delta_multiplier
        load_target_df.loc[load_target_df['Include']==True, 'DistMvarInput'] += load_target_df['DistMvarInput_Delta'] * delta_multiplier
        return
    
    def create_statcom_on_lowestv_bus(vpu_min = 0.85, vnom_min = 50) -> int:
        SimAuto.SaveState()
        # If there is a bus with voltage lower than v_min, this will 
        bus_params: dict[str,type] = {
            'Number': int
            ,'BusNomVolt': float
            ,'Vpu': float
            ,'IslandNumber': int
            ,'BusIsStarBus:1': str
        }
        bus_df = get_param_df(SimAuto, 'Bus', bus_params)
        filtered_df = bus_df[(bus_df['IslandNumber'] == 1) & 
                             (bus_df['BusIsStarBus:1'] == 'NO') & 
                             (bus_df['BusNomVolt'] > vnom_min) &
                             (bus_df['Vpu'] < vpu_min) ]
        if len(filtered_df) == 0:
            return 0

        min_row = filtered_df.loc[filtered_df['Vpu'].idxmin()]
        vpu = min_row['Vpu']
        number = min_row['Number']
        nomvolt = min_row['BusNomVolt']

        print(f'Adding statcom to: {number}, kV={nomvolt}, Vpu={vpu}')
        # Create a STATCOM, in the "Open" position. Solve case. 
        statcom_dict = {
            'BusNum': number
            ,'GenID': 'xx'
            ,'GenStatus': 'Open'
            ,'GenAVRAble': 'YES'
            ,'GenMVRMax': 9999
            ,'GenMVRMin': -9999
            ,'GenMvrSetPoint': 0
            ,'GenAGCAble': 'NO'
            ,'GenMWMax': 0
            ,'GenMWMin': 0
            ,'GenMWSetPoint': 0
            ,'GenVoltSet': 1.000   
        }
        statcom_df = pd.DataFrame([statcom_dict])

        SimAuto.RunScriptCommand('EnterMode(EDIT);')
        SimAuto.CreateIfNotFound = True
        set_param_df(SimAuto, 'Gen', statcom_df)
        SimAuto.RunScriptCommand('EnterMode(RUN);')

        if solve(SimAuto):
            SimAuto.SaveState()
        else:
            print('WARNING: Did not solve after creating statcom in Open state. Rolling back change.')
            SimAuto.LoadState()
            return number

        # Try to close STATCOM. Roll back if it doesn't solve.
        SimAuto.SaveState()
        statcom_dict['GenStatus'] = 'Closed'
        set_param_df(SimAuto, 'Gen', statcom_df)
        if not solve(SimAuto, mva_mismatch_threshold):
            print('WARNING: Did not solve after closing statcom. Rolling back change.')
            SimAuto.LoadState()

        return number

    def drop_collapsed_sections(vpu_min = 0.80, vpu_max = 0.80):
        SimAuto.SaveState()

        # Disconnects network sections which are beginning to show collapse. 
        # I.e. opens branches connecting from buses with OK voltage (>vpu_max) to 
        # buses with awful voltage (<vpu_min).
        branch_params: dict[str,type] = {
            'ObjectID': str
            ,'Status': str
            ,'BranchVpuHigh': float
            ,'BranchVpuLow': float
        }
        df = get_param_df(SimAuto, 'Branch', branch_params)
        filtered_df = df[(df['Status'] == 'Closed') & 
                             (df['BranchVpuHigh'] > vpu_max) & 
                             (df['BranchVpuLow'] < vpu_min)].copy()
        if len(filtered_df) == 0:
            return set([0])
        
        print(f'Dropping branches to collapsed network sections: {filtered_df["ObjectID"].unique()}')
        filtered_df['Status'] = 'Open'
        set_param_df(SimAuto, 'Branch', filtered_df)
        # Run command "ClearSmallIslands;"
        SimAuto.RunScriptCommand('ClearSmallIslands;')

        if not solve(SimAuto, mva_mismatch_threshold):
            print('WARNING: Did not solve after dropping branches. Rolling back change.')
            SimAuto.LoadState()
        return set(filtered_df['ObjectID'].unique())

    # Save state. 
    SimAuto.SaveState()

    # Setup logs. 
    scalelog_dict: dict[str,pd.DataFrame] = {}
    statcom_bus_set = set()
    dropped_branch_set = set()

    if not solve(SimAuto, mva_mismatch_threshold):
        SimAuto.LoadState()
        print('Could not solve the original input case!')
        scalelog_dict['iteration_df'] = pd.DataFrame({"Value": ['Failed to converge base case.']})
        return scalelog_dict

    [gen_pvqv_df, load_pvqv_df] = compute_pvqv_exclusions()
    compute_voltage_exclusions()
    close_all_related_gen_load()
    compute_deltas()

    adjust_shunts(SimAuto)
    if not solve(SimAuto):
        SimAuto.LoadState()
        print('Could not solve after adjusting shunts in the original case!')
        scalelog_dict['iteration_df'] = pd.DataFrame({"Value": ['Failed to converge base case with shunt adjustments.']})
        return scalelog_dict

    print('')
    iteration_success = True
    for iteration in range(iterations):
        SimAuto.SaveState()
        print(f'\r----- Iteration: {iteration} of {iterations} -----           ') # , end='')
        increment(1.0)
        set_param_df(SimAuto, 'Gen', gen_target_df)
        set_param_df(SimAuto, 'Load', load_target_df)
        if solve(SimAuto) and solve(SimAuto):
            SimAuto.SaveState()
            adjust_shunts(SimAuto)
            compute_voltage_exclusions()
            dropped_branches = drop_collapsed_sections()
            dropped_branch_set.update(dropped_branches)
            statcom_number = create_statcom_on_lowestv_bus()
            statcom_bus_set.add(statcom_number)
        else:
            print(f'Stopped at Iteration: {iteration} of {iterations}')
            print('Iteration did not solve. Reverting iteration and stopping.')
            increment(-1.0)
            SimAuto.LoadState() # SimAuto.RunScriptCommand("RestoreState('LASTSUCCESSFUL','');")
            iteration_success = False
            break # Exit the for-loop.

    # Package the logs for return. 
    scalelog_dict['gen'] = gen_target_df[gen_target_df['Include'] == False]
    scalelog_dict['load'] = load_target_df[load_target_df['Include'] == False]
    scalelog_dict['iteration_df'] = pd.DataFrame({"Value": [iteration]})
    scalelog_dict['statcom_bus_df'] = pd.DataFrame({"BusNum": list(statcom_bus_set)})
    scalelog_dict['gen_pvqv'] = gen_pvqv_df
    scalelog_dict['load_pvqv'] = load_pvqv_df
    scalelog_dict['dropped_branch_df'] = pd.DataFrame({"ObjectID": list(dropped_branch_set)})
    
    if not iteration_success:
        return scalelog_dict
    
    compute_voltage_exclusions()
    SimAuto.SaveState()
    gen_final_status_change_df = set_gen_load_status()
    if not solve(SimAuto):
        print('Setting final gen/load statuses did not succeed. Reverting change.')
        SimAuto.LoadState()

    # Package the logs for return. 
    scalelog_dict['gen_final_st_change'] = gen_final_status_change_df[gen_final_status_change_df['FinalStatusChange'] == 'DIVERGED']

    print('Reached end of iterate_to_gen_load_targets(). Returning.')

    return scalelog_dict

def set_branch_statuses(SimAuto, left_case_dict, right_case_dict):
    """
    With the "Right" case open, sets the "Right" model branch statuses to match those from the "Left" model. 
    """
    
    # Get Transformer and Non-Transformer branch statuses by ObjectID. 
    left_branch_df = pd.concat([ 
        left_case_dict['Branch']['df'][['ObjectID','Status','BranchDeviceType']]
        ,left_case_dict['Transformer']['df'][['ObjectID','Status','BranchDeviceType']]
        ])
    right_branch_df = pd.concat([ 
        right_case_dict['Branch']['df'][['ObjectID','Status','BranchDeviceType']]
        ,right_case_dict['Transformer']['df'][['ObjectID','Status','BranchDeviceType']]
        ])
    
    # GridView does not export Breakers and Disconnects in EPCs. 
    # Exclude Breakers and Disconnects from the status changes. 
    left_branch_df = left_branch_df[~left_branch_df['BranchDeviceType'].isin(['Breaker', 'Disconnect'])]
    left_branch_df.drop(columns=['BranchDeviceType'], inplace=True)

    right_branch_df = right_branch_df[~right_branch_df['BranchDeviceType'].isin(['Breaker', 'Disconnect'])]
    right_branch_df.drop(columns=['BranchDeviceType'], inplace=True)

    # right df, with ['ObjectID', 'Status', 'StatusLeft']
    merged_df = right_branch_df.merge(
        left_branch_df
        ,on='ObjectID'
        ,how='left'
        ,suffixes=('', 'Left')
    )

    # Save a copy of the status targets for reporting purposes. 
    status_targets_df = merged_df.copy()
    status_targets_df = status_targets_df[status_targets_df['Status'] != status_targets_df['StatusLeft']]
    status_targets_df.sort_values(by='StatusLeft', inplace=True)

    # Any branches that don't exist on the left will need to be opened. 
    merged_df['StatusLeft'] = merged_df['StatusLeft'].fillna('Open')

    # Branches which need statuses modified. 
    filtered_df = merged_df[merged_df['Status'] != merged_df['StatusLeft']].copy(deep=True)
    
    # Save the current state. 
    filtered_df['StatusRight'] = filtered_df['Status']

    # Empty df to keep track of any branches which couldn't solve. 
    fail_df = pd.DataFrame(columns=filtered_df.columns)

    # Attempt to solve all changes at once.
    print(f'Attempting to set statuses on all branches at the same time. ')
    SimAuto.SaveState()
    allchanges_df = filtered_df.copy(deep=True)
    allchanges_df['Status'] = allchanges_df['StatusLeft']
    message = set_param_df(SimAuto, 'Branch', allchanges_df)
    if not solve(SimAuto, mva_mismatch_threshold):
        # Failed to do all changes at once! Revert, and try individual branch changes. 
        print(f'Failed to set status on all elements at the same time! Testing individual branches one at a time.')
        SimAuto.LoadState()

        # For each object_id, attempt to change the status and solve. 
        # If it fails, keep track of those failures. 
        counter = 1
        total_count = len(filtered_df['ObjectID'].unique())
        for object_id in filtered_df['ObjectID'].unique():
            print(f'Item {counter} of {total_count}: Setting status on {object_id}')
            counter += 1
            row_df = filtered_df[filtered_df['ObjectID']==object_id].copy(deep=True)

            # Save state. Attempt modification. Solve. Revert and log if failed to solve. 
            SimAuto.SaveState()
            row_df['Status'] = row_df['StatusLeft']
            message = set_param_df(SimAuto, 'Branch', row_df)
            if not solve(SimAuto, mva_mismatch_threshold):
                SimAuto.LoadState()
                fail_df = pd.concat([fail_df, row_df], ignore_index = True)
                print(f'Failed to set status on {object_id}')

    return [status_targets_df, fail_df]

def adjust_shunts(SimAuto, vlow: float = 0.92, vhigh: float = 1.08, max_iterations: int = 10):
    # Save state.
    SimAuto.SaveState()

    # Adjusts shunts to attempt to get buses back within a set voltage band. 

    table = 'Shunt'
    parameter_type: dict[str,type] = {
        'ObjectID': str
        ,'MvarNom': float
        ,'Status': str
        ,'Vpu': float
        ,'IslandNumber': int
    }

    def suggested_shunt_status(row):
        if(row['MvarNom'] < 0): # Reactor
            if(row['Status'] == 'Closed' and row['Vpu'] < vlow):
                return 'Open'
            if(row['Status'] == 'Open' and row['Vpu'] > vhigh):
                return 'Closed'
        if(row['MvarNom'] > 0): # Capacitor
            if(row['Status'] == 'Open' and row['Vpu'] < vlow):
                return 'Closed'
            if(row['Status'] == 'Closed' and row['Vpu'] > vhigh):
                return 'Open'
        return row['Status'] # Keep as-is. 
    
    def get_suggested_statuses():
        df = get_param_df(SimAuto, table, parameter_type)
        df['NewStatus'] = df.apply(suggested_shunt_status, axis=1)
        change_df = df[df['Status']!=df['NewStatus']].copy(deep=True)
        return change_df

    def adjust_all_shunts():
        # Adjusts all shunts in one go, to get the bulk of the work done. 
        SimAuto.SaveState()

        df = get_suggested_statuses()
        df['Status'] = df['NewStatus']
        message = set_param_df(SimAuto, table, df)

        if not solve(SimAuto, mva_mismatch_threshold):
            SimAuto.LoadState()

        return

    def iterate_on_individual_shunts():
        # Adjusts 1 shunt, solves, sees if any more need adjustment.
        # This helps fix situations where a bus has several shunts, and only some need to be online. 
        for index in range(max_iterations):
            SimAuto.SaveState()

            df = get_suggested_statuses()
            df['Status'] = df['NewStatus']
            # Adjust the smallest shunts first
            df = df.sort_values(by='MvarNom', key=abs)
            if(len(df) == 0):
                return # Nothing left to do! All shunt voltages within bounds. 
            message = set_param_df(SimAuto, table, df.iloc[[0]])

            if not solve(SimAuto, mva_mismatch_threshold):
                SimAuto.LoadState()

        return
    adjust_all_shunts()
    iterate_on_individual_shunts()
    if not solve(SimAuto, mva_mismatch_threshold):
        print('adjust_shunts() did not solve. Restoring state.')
        SimAuto.LoadState()
    return

def fix_transformer_taps(SimAuto, threshold = 0.15):
    # In GridView, some transformer models may have transformer nominal kV and tap positions which are aggregious errors.
    # GridView doesn't have issues with this (DC Loadflow), but this certainly will cause problems in PowerWorld. 
    # If Nominal kV and Tap positions are far out of range, fix them to 1.0 PU based on the bus nominal kV.

    def is_transformer_okay(row):
        def is_variable_tap_okay(row):
            return abs(row['Tap'] - 1.0) < threshold
        def is_from_okay(row):
            return (abs((row['NomkVFrom'] / row['XFNomkVbaseFrom']) - 1.0) < threshold) and ((row['TapFixedFrom'] - 1.0) < threshold)
        def is_to_okay(row):
            return (abs((row['NomkVTo'] / row['XFNomkVbaseTo']) - 1.0) < threshold) and ((row['TapFixedTo'] - 1.0) < threshold)
        ret_val = is_variable_tap_okay(row) and is_from_okay(row) and is_to_okay(row)
        return ret_val

    table = 'Branch'
    parameter_type: dict[str,type] = {
        'ObjectID': str
        ,'ControlType': str
        ,'NomkVFrom': float
        ,'XFNomkVbaseFrom': float
        ,'NomkVTo': float
        ,'XFNomkVbaseTo': float
        ,'TapFixedFrom': float
        ,'TapFixedTo': float
        ,'Tap': float
    }
    df = get_param_df(SimAuto, 'Branch', parameter_type, "BranchDeviceType = 'Transformer'")

    df['okay'] = df.apply(is_transformer_okay, axis=1)
    bad_df = df[df['okay']==False].copy(deep=True)

    fix_df = bad_df.copy(deep=True)
    fix_df['XFNomkVbaseFrom'] = fix_df['NomkVFrom']
    fix_df['XFNomkVbaseTo'] = fix_df['NomkVTo']
    fix_df['TapFixedFrom'] = 1.0
    fix_df['TapFixedTo'] = 1.0
    fix_df['Tap'] = 1.0

    message = set_param_df(SimAuto, table, fix_df)
    
    return bad_df

def get_fault_duty(SimAuto) -> pd.DataFrame:
    """
    Returns a dataframe of fault duty at every bus. 
    """
    # Get a list of all buses. 
    bus_params: dict[str,type] = {
        'ObjectID': str
        ,'Number': int
        ,'Name': str
        ,'NomkV': float
    }
    bus_df = get_param_df(SimAuto, 'Bus', bus_params)

    fault_df = pd.DataFrame()
    fault_df['FaultName'] = bus_df['Number'].astype(str) + ' : ' + bus_df['Name'].astype(str) + ' : ' + bus_df['NomkV'].astype(str)
    fault_df['WhoAmI'] = bus_df['ObjectID']
    fault_df['CustomInteger:0'] = bus_df['Number']
    fault_df['FaultType'] = '3PB'
    fault_df['FaultImpedance'] = 0
    fault_df['FaultImpedance:1'] = 0
    
    SimAuto.RunScriptCommand('EnterMode(EDIT);')
    SimAuto.CreateIfNotFound = True
    SimAuto.RunScriptCommand('Delete(Fault);')
    set_param_df(SimAuto, 'Fault', fault_df)
    # set_param_df(SimAuto, 'Fault', fault_df.head(10)) # Test with only the top 10.

    fault_params = {
        'FaultName': str
        ,'WhoAmI': str
        ,'CustomInteger:0': int
        ,'FaultType': str
        ,'FaultImpedance': float
        ,'FaultImpedance:1': float
        ,'BusNomVolt': float
        ,'ABCPhaseI': float
    }

    SimAuto.RunScriptCommand('EnterMode(RUN);')
    SimAuto.RunScriptCommand('FaultMultiple(NO);')
    
    result_df = get_param_df(SimAuto, 'Fault', fault_params)
    result_df.rename(columns={"CustomInteger:0": "BusNumber"}, inplace=True)
    # MVA = sqrt(3) * kV * kA
    result_df['MVA'] = (3 ** 0.5) * result_df['BusNomVolt'] * result_df['ABCPhaseI']

    return result_df

def get_pvqv(SimAuto) -> pd.DataFrame:
    """
    Calculates bus voltage MW & MVAR self-sensitivity. 
    Tools -> Sensitivities -> Flow and Voltage Sensitivities -> Self Sensitivity (Tab)
    """
    bus_params = {
        'Number': int
        ,'Name': str
        ,'NomkV': float
        ,'AreaNumber': int
        ,'AreaName': str
        ,'SensdVdPself': float
        ,'SensdVdQself': float
        ,'Vpu': float
    }

    SimAuto.RunScriptCommand('EnterMode(RUN);')
    SimAuto.RunScriptCommand('CalculateVoltSelfSense();')
    result_df = get_param_df(SimAuto, 'Bus', bus_params)
    return result_df

def auto_fit_columns(writer: pd.ExcelWriter):
    workbook = writer.book
    for sheet_name in writer.sheets:
        worksheet = workbook[sheet_name]
        for col in worksheet.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = max_length + 2

def freeze_top_rows(writer: pd.ExcelWriter):
    for sheet in writer.sheets.values():
        sheet.freeze_panes = 'A2'  # Freeze the top row
    return

def filter_top_rows(writer: pd.ExcelWriter):
    for sheet in writer.sheets.values():
        max_column = sheet.max_column
        max_column_letter = openpyxl.utils.get_column_letter(max_column)
        sheet.auto_filter.ref = f"A1:{max_column_letter}1"  # Adjust range based on columns
    return

def df_dict_to_excel_workbook(rep_fp: Path, dict_df: dict[str,pd.DataFrame]):
    """Writes a dictionary of dataframes to an Excel Workbook."""
    writer = pd.ExcelWriter(rep_fp, engine='openpyxl')
    for sheet_name, df in dict_df.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    auto_fit_columns(writer)
    freeze_top_rows(writer)
    filter_top_rows(writer)
    try:
        writer.close()
    except:
        pass
